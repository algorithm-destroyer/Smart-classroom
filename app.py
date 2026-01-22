from flask import Flask, request, jsonify, render_template, redirect, url_for, session, send_file
from firebase_admin import credentials, firestore, initialize_app, auth
import datetime
import os
import requests
from functools import wraps
import base64
import cv2
import face_recognition
import numpy as np
from openpyxl import Workbook
import uuid
import shutil
import logging
import random

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.urandom(24)  # Secure random secret key for session

# Firebase Web API Key for REST calls
API_KEY = "AIzaSyAkPeCdgxk1TyM8KKFjZc-ebbEt1KK41tU"

# Initialize Firebase
try:
    if not os.path.exists('serviceAccountKey.json'):
        raise FileNotFoundError("serviceAccountKey.json not found in project root")
    cred = credentials.Certificate('serviceAccountKey.json')
    firebase_app = initialize_app(cred)
    db = firestore.client()  # Firestore client
except Exception as e:
    logger.error(f"Error initializing Firebase: {str(e)}")
    raise

# Create temp folder for Excel storage
TEMP_DIR = os.path.join(os.getcwd(), 'temp')
if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR)

# Decorator to check if user is logged in (for students)
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# Decorator to check if teacher is logged in
def teacher_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'teacher_id' not in session:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# Helper function to get face encoding from base64 image
def get_face_encoding(image_base64):
    try:
        if ',' in image_base64:
            image_base64 = image_base64.split(',')[1]
        image_data = base64.b64decode(image_base64)
        nparr = np.frombuffer(image_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None:
            raise ValueError("Invalid image data")
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        locations = face_recognition.face_locations(rgb_img)
        if len(locations) == 0:
            raise ValueError("No face detected. Try again.")
        if len(locations) > 1:
            raise ValueError("Multiple faces detected. Please capture only one face.")
        encodings = face_recognition.face_encodings(rgb_img, locations)
        return encodings[0]
    except Exception as e:
        raise e

# Route for the intro page (first page)
@app.route('/')
def index():
    return render_template('intro.html')

# Route for the entry page
@app.route('/entry')
def entry_page():
    return render_template('entry_page.html')

# Route for the student login page
@app.route('/login')
def login():
    return render_template('login.html')

# Route for the teacher login page
@app.route('/teacher_login')
def teacher_login():
    return render_template('teacher_login.html')

# Route for the registration page
@app.route('/register')
def register_page():
    return render_template('register.html')

# Route for the teacher registration page
@app.route('/teacher_register')
def teacher_register_page():
    return render_template('teacher_register.html')

# Route for the forgot page
@app.route('/forgot')
def forgot_page():
    return render_template('forgot.html')

# Route for the student home page
@app.route('/home')
@login_required
def home():
    try:
        user_id = session.get('user_id')
        if not user_id:
            logger.warning("No user_id in session, redirecting to index")
            return redirect(url_for('index'))

        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            logger.error(f"User data not found for user_id: {user_id}")
            return jsonify({
                'status': 'error',
                'message': 'User data not found'
            }), 404

        user_data = user_doc.to_dict()
        logger.debug(f"Fetched user_data for home: {user_data}")
        
        user_data['name'] = user_data.get('name', 'Unknown Student')
        user_data['regNo'] = user_data.get('regNo', 'Unknown RegNo')
        user_data['email'] = user_data.get('email', 'Not available')
        user_data['branch'] = user_data.get('branch', 'Not available')
        user_data['year'] = user_data.get('year', 'Not available')
        user_data['phone'] = user_data.get('phone', 'Not available')
        user_data['dob'] = user_data.get('dob', 'Not available')
        user_data['gender'] = user_data.get('gender', 'Not available')
        
        if 'face_image' in user_data and user_data['face_image']:
            user_data['profile_photo'] = f"data:image/jpeg;base64,{user_data['face_image']}"
        else:
            user_data['profile_photo'] = 'https://picsum.photos/seed/student123/180/180.jpg'

        return render_template('home.html', user_data=user_data)
        
    except Exception as e:
        logger.error(f"Error fetching user data for home: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for attendance page
@app.route('/attendance.html')
@login_required
def attendance():
    return render_template('attendance.html')

# Route for assignments page
@app.route('/assignments.html')
@login_required
def assignments():
    return render_template('assignments.html')

# Route for recorded classes page
@app.route('/recorded_class.html')
@login_required
def recorded_classes():
    return render_template('recorded_class.html')

# Route for notes page
@app.route('/notes.html')
@login_required
def notes():
    return render_template('notes.html')

# Route for notices page
@app.route('/notices.html')
@login_required
def notices():
    return render_template('notices.html')

# Route for profile page
@app.route('/profile.html')
@login_required
def profile():
    try:
        user_id = session.get('user_id')
        if not user_id:
            logger.warning("No user_id in session, redirecting to index")
            return redirect(url_for('index'))

        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            logger.error(f"User data not found for user_id: {user_id}")
            return jsonify({
                'status': 'error',
                'message': 'User data not found'
            }), 404

        user_data = user_doc.to_dict()
        logger.debug(f"Fetched user_data: {user_data}")
        
        if 'face_image' in user_data and user_data['face_image']:
            user_data['profile_photo'] = f"data:image/jpeg;base64,{user_data['face_image']}"
        else:
            user_data['profile_photo'] = 'https://picsum.photos/seed/student123/180/180.jpg'

        user_data['address'] = user_data.get('address', 'Not available')
        user_data['school'] = user_data.get('school', 'Not available')
        user_data['schoolAddress'] = user_data.get('schoolAddress', 'Not available')
        user_data['schoolId'] = user_data.get('schoolId', 'Not available')
        user_data['classroom'] = user_data.get('classroom', 'Not available')
        user_data['subjects'] = user_data.get('subjects', ['No subjects'])
        user_data['name'] = user_data.get('name', 'Unknown Student')
        user_data['regNo'] = user_data.get('regNo', 'Unknown RegNo')

        return render_template('profile.html', user_data=user_data)
        
    except Exception as e:
        logger.error(f"Error fetching user data: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for settings page
@app.route('/settings.html')
@login_required
def settings():
    try:
        user_id = session.get('user_id')
        if not user_id:
            logger.warning("No user_id in session, redirecting to index")
            return redirect(url_for('index'))

        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            logger.error(f"User data not found for user_id: {user_id}")
            return jsonify({
                'status': 'error',
                'message': 'User data not found'
            }), 404

        user_data = user_doc.to_dict()
        logger.debug(f"Fetched user_data for settings: {user_data}")
        
        user_data['name'] = user_data.get('name', 'Unknown Student')
        user_data['regNo'] = user_data.get('regNo', 'Unknown RegNo')

        return render_template('settings.html', user_data=user_data)
        
    except Exception as e:
        logger.error(f"Error fetching user data for settings: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for academic profile page
@app.route('/academic_profile.html')
@login_required
def academic_profile():
    try:
        user_id = session.get('user_id')
        if not user_id:
            logger.warning("No user_id in session, redirecting to index")
            return redirect(url_for('index'))

        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            logger.error(f"User data not found for user_id: {user_id}")
            return jsonify({
                'status': 'error',
                'message': 'User data not found'
            }), 404

        user_data = user_doc.to_dict()
        logger.debug(f"Fetched user_data for academic_profile - name: {user_data.get('name')}, regNo: {user_data.get('regNo')}")

        academic_data = {
            'name': user_data.get('name', 'Unknown Student'),
            'regNo': user_data.get('regNo', 'Unknown RegNo')
        }

        return render_template('academic_profile.html', user_data=academic_data)
        
    except Exception as e:
        logger.error(f"Error fetching user data for academic_profile: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher home page
@app.route('/teacher_home.html')
@teacher_login_required
def teacher_home():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for home: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')

        return render_template('teacher_home.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for home: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher settings page
@app.route('/teacher_settings.html')
@teacher_login_required
def teacher_settings():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for settings: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')
        teacher_data['gender'] = teacher_data.get('gender', 'Not available')
        teacher_data['dob'] = teacher_data.get('dob', 'Not available')

        return render_template('teacher_settings.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for settings: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher faculty advisor page
@app.route('/teacher_fa.html')
@teacher_login_required
def teacher_fa():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        # Fetch teacher data
        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_fa: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')

        # Fetch all students
        users = db.collection('users').stream()
        students = []
        for user in users:
            data = user.to_dict()
            students.append({
                'name': data.get('name', 'Unknown'),
                'regNo': data.get('regNo', 'Not available'),
                'email': data.get('email', 'Not available'),
                'phone': data.get('phone', 'Not available'),
                'gender': data.get('gender', 'Not available'),
                'dob': data.get('dob', 'Not available'),
                'branch': data.get('branch', 'Not available')
            })
        students.sort(key=lambda x: x['regNo'])
        logger.debug(f"Fetched {len(students)} students for teacher_fa")

        return render_template('teacher_fa.html', teacher_data=teacher_data, students=students)
        
    except Exception as e:
        logger.error(f"Error fetching data for teacher_fa: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher assignment page
@app.route('/teacher_assignment.html')
@teacher_login_required
def teacher_assignment():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_assignment: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')

        return render_template('teacher_assignment.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_assignment: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher notice page
@app.route('/teacher_notice.html')
@teacher_login_required
def teacher_notice():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_notice: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')

        return render_template('teacher_notice.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_notice: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher recorded class page
@app.route('/teacher_recorded_class.html')
@teacher_login_required
def teacher_recorded_class():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_recorded_class: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')

        return render_template('teacher_recorded_class.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_recorded_class: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher query page
@app.route('/teacher_query.html')
@teacher_login_required
def teacher_query():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_query: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')

        return render_template('teacher_query.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_query: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher study material page
@app.route('/teacher_study.html')
@teacher_login_required
def teacher_study():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_study: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')

        return render_template('teacher_study.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_study: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher analytics dashboard page
@app.route('/teacher_analytics.html')
@teacher_login_required
def teacher_analytics():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_analytics: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')

        # Fetch all students from Firestore
        users = db.collection('users').stream()
        students = []
        classes = ["Class 10A", "Class 10B", "Class 11A", "Class 11B"]
        for user in users:
            data = user.to_dict()
            avg_score = random.randint(50, 95)
            performance = (
                "Excellent" if avg_score >= 90 else
                "Good" if avg_score >= 75 else
                "Average" if avg_score >= 60 else
                "Poor"
            )
            students.append({
                'regNo': data.get('regNo', 'Not available'),
                'name': data.get('name', 'Unknown'),
                'class': random.choice(classes),
                'attendance': f"{random.randint(70, 95)}%",
                'avg_score': f"{avg_score}%",
                'performance': performance
            })
        students.sort(key=lambda x: x['regNo'])
        logger.debug(f"Fetched {len(students)} students for teacher_analytics")

        return render_template('teacher_analytics.html', teacher_data=teacher_data, students=students)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_analytics: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher attendance management page
@app.route('/teacher_attendance.html')
@teacher_login_required
def teacher_attendance():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_attendance: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')

        return render_template('teacher_attendance.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_attendance: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher profile page
@app.route('/teacher_profile.html')
@teacher_login_required
def teacher_profile():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_profile: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')
        teacher_data['gender'] = teacher_data.get('gender', 'Not available')
        teacher_data['dob'] = teacher_data.get('dob', 'Not available')

        return render_template('teacher_profile.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_profile: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# Route for teacher recorded notes page
@app.route('/teacher_notes.html')
@teacher_login_required
def teacher_notes():
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            logger.warning("No teacher_id in session, redirecting to index")
            return redirect(url_for('index'))

        teacher_ref = db.collection('teachers').document(teacher_id)
        teacher_doc = teacher_ref.get()
        
        if not teacher_doc.exists:
            logger.error(f"Teacher data not found for teacher_id: {teacher_id}")
            return jsonify({
                'status': 'error',
                'message': 'Teacher data not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        logger.debug(f"Fetched teacher_data for teacher_notes: {teacher_data}")
        
        teacher_data['name'] = teacher_data.get('name', 'Unknown Teacher')
        teacher_data['employeeId'] = teacher_data.get('employeeId', 'Unknown Employee ID')
        teacher_data['email'] = teacher_data.get('email', 'Not available')
        teacher_data['phone'] = teacher_data.get('phone', 'Not available')
        teacher_data['subject'] = teacher_data.get('subject', 'Not available')
        teacher_data['department'] = teacher_data.get('department', 'Not available')

        return render_template('teacher_notes.html', teacher_data=teacher_data)
        
    except Exception as e:
        logger.error(f"Error fetching teacher data for teacher_notes: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# API endpoint to generate report card as Excel
@app.route('/api/generate_report_card', methods=['POST'])
@login_required
def generate_report_card():
    try:
        user_id = session.get('user_id')
        if not user_id:
            logger.warning("No user_id in session for generate_report_card")
            return jsonify({
                'status': 'error',
                'message': 'User not logged in'
            }), 401

        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            logger.error(f"User data not found for user_id: {user_id}")
            return jsonify({
                'status': 'error',
                'message': 'User data not found'
            }), 404

        user_data = user_doc.to_dict()
        student_name = user_data.get('name', 'Unknown Student')
        reg_no = user_data.get('regNo', 'Unknown RegNo')
        logger.debug(f"User data for report card: {user_data}")

        data = request.get_json()
        exam_name = data.get('exam_name')
        if not exam_name:
            logger.error("Exam name missing in generate_report_card request")
            return jsonify({
                'status': 'error',
                'message': 'Exam name is required'
            }), 400

        exam_data = {
            'Mid-Term Examination': [
                {'subject': 'Mathematics', 'marks': 85},
                {'subject': 'Physics', 'marks': 78},
                {'subject': 'Computer Science', 'marks': 92},
                {'subject': 'English', 'marks': 88}
            ],
            'Final Examination': [
                {'subject': 'Mathematics', 'marks': 90},
                {'subject': 'Physics', 'marks': 82},
                {'subject': 'Computer Science', 'marks': 95},
                {'subject': 'English', 'marks': 85}
            ],
            'Practical Assessment': [
                {'subject': 'Lab Work', 'marks': 88},
                {'subject': 'Project', 'marks': 90}
            ],
            'Quiz 1: Data Structures': [
                {'subject': 'Data Structures', 'marks': 80}
            ]
        }

        if exam_name not in exam_data:
            logger.error(f"Invalid exam name: {exam_name}")
            return jsonify({
                'status': 'error',
                'message': 'Invalid exam name'
            }), 400

        marks = exam_data[exam_name]

        wb = Workbook()
        ws = wb.active
        ws.title = "Report Card"

        ws['A1'] = 'Report Card'
        ws['A2'] = 'Student Name'
        ws['B2'] = student_name
        ws['A3'] = 'Registration Number'
        ws['B3'] = reg_no
        ws['A4'] = 'Examination'
        ws['B4'] = exam_name
        ws['A5'] = 'Date'
        ws['B5'] = datetime.date.today().strftime('%B %d, %Y')

        ws['A7'] = 'Subject'
        ws['B7'] = 'Marks (Out of 100)'

        for i, subject in enumerate(marks, start=8):
            ws[f'A{i}'] = subject['subject']
            ws[f'B{i}'] = subject['marks']

        for column in ['A', 'B']:
            max_length = 0
            column_letter = column
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        excel_filename = f"{exam_name.replace(' ', '_')}_Report_Card_{uuid.uuid4().hex}.xlsx"
        excel_path = os.path.join(TEMP_DIR, excel_filename)
        wb.save(excel_path)
        logger.debug(f"Generated Excel: {excel_path}")

        response = send_file(
            excel_path,
            as_attachment=True,
            download_name=f"{exam_name} Report Card.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        @response.call_on_close
        def cleanup():
            try:
                if os.path.exists(excel_path):
                    os.remove(excel_path)
                    logger.debug(f"Cleaned up Excel: {excel_path}")
            except Exception as e:
                logger.error(f"Error cleaning up Excel: {str(e)}")

        return response

    except Exception as e:
        logger.error(f"Error generating report card: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# API endpoint for user registration
@app.route('/api/register', methods=['POST'])
def register_user():
    try:
        data = request.get_json()
        
        required_fields = ['name', 'gender', 'dob', 'branch', 'year', 'email', 'regNo', 'phone', 'password']
        for field in required_fields:
            if field not in data or not data[field]:
                logger.error(f"Missing required field: {field}")
                return jsonify({
                    'status': 'error',
                    'message': f'Missing required field: {field}'
                }), 400
        
        if not data['name'].replace(' ', '').isalpha():
            logger.error("Invalid name format")
            return jsonify({
                'status': 'error',
                'message': 'Name should only contain letters and spaces'
            }), 400
            
        if not data['branch'].replace(' ', '').isalpha():
            logger.error("Invalid branch format")
            return jsonify({
                'status': 'error',
                'message': 'Branch should only contain letters and spaces'
            }), 400
            
        if data['year'] not in ['1st Year', '2nd Year', '3rd Year', '4th Year', 'PG 1st Year', 'PG 2nd Year']:
            logger.error(f"Invalid year: {data['year']}")
            return jsonify({
                'status': 'error',
                'message': 'Invalid year selected'
            }), 400
            
        if '@' not in data['email']:
            logger.error("Invalid email format")
            return jsonify({
                'status': 'error',
                'message': 'Invalid email format'
            }), 400
            
        if not data['phone'].isdigit() or len(data['phone']) != 10:
            logger.error("Invalid phone number")
            return jsonify({
                'status': 'error',
                'message': 'Phone number must be 10 digits'
            }), 400
            
        try:
            dob = datetime.datetime.strptime(data['dob'], '%Y-%m-%d')
            today = datetime.datetime.now()
            min_date = today - datetime.timedelta(days=100*365)
            if dob > today or dob < min_date:
                logger.error("Invalid date of birth")
                return jsonify({
                    'status': 'error',
                    'message': 'Date of birth must not be in the future or more than 100 years ago'
                }), 400
        except ValueError:
            logger.error("Invalid date of birth format")
            return jsonify({
                'status': 'error',
                'message': 'Invalid date of birth format'
            }), 400
            
        users_ref = db.collection('users')
        reg_check = users_ref.where('regNo', '==', data['regNo']).limit(1).stream()
        if any(reg_check):
            logger.error(f"Registration number already exists: {data['regNo']}")
            return jsonify({
                'status': 'error',
                'message': 'Registration number already exists'
            }), 409
            
        user = auth.create_user(
            email=data['email'],
            password=data['password']
        )
        uid = user.uid
        logger.debug(f"Created Firebase Auth user: {uid}")
        
        user_data = {
            'name': data['name'],
            'gender': data['gender'],
            'dob': data['dob'],
            'branch': data['branch'],
            'year': data['year'],
            'email': data['email'],
            'regNo': data['regNo'],
            'phone': data['phone'],
            'role': 'student',
            'created_at': datetime.datetime.now(),
            'updated_at': datetime.datetime.now()
        }
        
        users_ref.document(uid).set(user_data)
        logger.debug(f"Stored user data in Firestore: {user_data}")
        
        return jsonify({
            'status': 'success',
            'message': 'Registration successful! Redirecting to login...'
        }), 200
        
    except Exception as e:
        logger.error(f"Error during registration: {str(e)}")
        error_message = str(e)
        if 'EMAIL_EXISTS' in error_message:
            error_message = 'Email already registered'
            status_code = 409
        elif 'WEAK_PASSWORD' in error_message:
            error_message = 'Password is too weak'
            status_code = 400
        else:
            error_message = f'An error occurred: {error_message}'
            status_code = 500
        return jsonify({
            'status': 'error',
            'message': error_message
        }), status_code

# API endpoint for teacher registration
@app.route('/api/teacher_register', methods=['POST'])
def register_teacher():
    try:
        data = request.get_json()
        
        required_fields = ['name', 'gender', 'dob', 'email', 'employeeId', 'department', 'subject', 'phone', 'password']
        for field in required_fields:
            if field not in data or not data[field]:
                logger.error(f"Missing required field: {field}")
                return jsonify({
                    'status': 'error',
                    'message': f'Missing required field: {field}'
                }), 400
        
        if not data['name'].replace(' ', '').isalpha():
            logger.error("Invalid name format")
            return jsonify({
                'status': 'error',
                'message': 'Name should only contain letters and spaces'
            }), 400
            
        if '@' not in data['email']:
            logger.error("Invalid email format")
            return jsonify({
                'status': 'error',
                'message': 'Invalid email format'
            }), 400
            
        if not data['phone'].isdigit() or len(data['phone']) != 10:
            logger.error("Invalid phone number")
            return jsonify({
                'status': 'error',
                'message': 'Phone number must be 10 digits'
            }), 400
            
        try:
            dob = datetime.datetime.strptime(data['dob'], '%Y-%m-%d')
            today = datetime.datetime.now()
            min_date = today - datetime.timedelta(days=100*365)
            if dob > today or dob < min_date:
                logger.error("Invalid date of birth")
                return jsonify({
                    'status': 'error',
                    'message': 'Date of birth must not be in the future or more than 100 years ago'
                }), 400
        except ValueError:
            logger.error("Invalid date of birth format")
            return jsonify({
                'status': 'error',
                'message': 'Invalid date of birth format'
            }), 400
            
        teachers_ref = db.collection('teachers')
        employee_id_check = teachers_ref.where('employeeId', '==', data['employeeId']).limit(1).stream()
        if any(employee_id_check):
            logger.error(f"Employee ID already exists: {data['employeeId']}")
            return jsonify({
                'status': 'error',
                'message': 'Employee ID already exists'
            }), 409
            
        user = auth.create_user(
            email=data['email'],
            password=data['password']
        )
        uid = user.uid
        logger.debug(f"Created Firebase Auth teacher: {uid}")
        
        # Set custom claim for teacher role
        auth.set_custom_user_claims(uid, {'role': 'teacher'})
        logger.debug(f"Set custom claim 'role: teacher' for user: {uid}")
        
        teacher_data = {
            'name': data['name'],
            'gender': data['gender'],
            'dob': data['dob'],
            'email': data['email'],
            'employeeId': data['employeeId'],
            'department': data['department'],
            'subject': data['subject'],
            'phone': data['phone'],
            'role': 'teacher',
            'created_at': datetime.datetime.now(),
            'updated_at': datetime.datetime.now()
        }
        
        teachers_ref.document(uid).set(teacher_data)
        logger.debug(f"Stored teacher data in Firestore: {teacher_data}")
        
        return jsonify({
            'status': 'success',
            'message': 'Teacher registration successful! Redirecting to login...'
        }), 200
        
    except Exception as e:
        logger.error(f"Error during teacher registration: {str(e)}")
        error_message = str(e)
        if 'EMAIL_EXISTS' in error_message:
            error_message = 'Email already registered'
            status_code = 409
        elif 'WEAK_PASSWORD' in error_message:
            error_message = 'Password is too weak'
            status_code = 400
        else:
            error_message = f'An error occurred: {error_message}'
            status_code = 500
        return jsonify({
            'status': 'error',
            'message': error_message
        }), status_code

# API endpoint for student login
@app.route('/api/login', methods=['POST'])
def login_user():
    try:
        data = request.get_json()
        
        if 'regNo' not in data or 'password' not in data:
            logger.error("Missing regNo or password in login request")
            return jsonify({
                'status': 'error',
                'message': 'Registration number and password are required'
            }), 400

        users_ref = db.collection('users')
        user_query = users_ref.where('regNo', '==', data['regNo']).limit(1).stream()
        user_doc = None
        for doc in user_query:
            user_doc = doc
            break

        if not user_doc:
            logger.error(f"Registration number not found: {data['regNo']}")
            return jsonify({
                'status': 'error',
                'message': 'Registration number not found'
            }), 404

        user_data = user_doc.to_dict()
        email = user_data['email']
        uid = user_doc.id
        logger.debug(f"Found user: {uid}, email: {email}")

        # Verify user exists in Firebase Auth
        try:
            user = auth.get_user(uid)
            # Check if user has teacher role
            if hasattr(user, 'custom_claims') and user.custom_claims and user.custom_claims.get('role') == 'teacher':
                logger.error(f"User {uid} attempted student login with teacher role")
                return jsonify({
                    'status': 'error',
                    'message': 'Please use teacher login portal'
                }), 403
        except auth.UserNotFoundError:
            logger.error(f"Firebase Auth user not found for user_id: {uid}")
            # Delete orphaned Firestore record
            users_ref.document(uid).delete()
            return jsonify({
                'status': 'error',
                'message': 'User account not found in authentication system. Please register again.'
            }), 404

        # Authenticate with Firebase
        url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={API_KEY}"
        payload = {
            "email": email,
            "password": data['password'],
            "returnSecureToken": True
        }
        response = requests.post(url, json=payload)
        
        if response.status_code == 200:
            session['user_id'] = uid
            logger.debug(f"Login successful for user_id: {uid}")
            return jsonify({
                'status': 'success',
                'message': 'Login successful! Redirecting to student portal...'
            }), 200
        else:
            error_info = response.json().get('error', {})
            message = error_info.get('message', 'Incorrect password')
            if 'INVALID_PASSWORD' in message:
                message = 'Incorrect password'
            elif 'USER_NOT_FOUND' in message:
                message = 'User not found'
            elif 'INVALID_LOGIN_CREDENTIALS' in message:
                message = 'Invalid login credentials'
            logger.error(f"Login failed: {message}")
            return jsonify({
                'status': 'error',
                'message': message
            }), 401

    except Exception as e:
        logger.error(f"Error during login: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# API endpoint for teacher login
@app.route('/api/teacher_login', methods=['POST'])
def teacher_login_user():
    try:
        data = request.get_json()
        
        if 'loginId' not in data or 'password' not in data:
            logger.error("Missing loginId or password in teacher login request")
            return jsonify({
                'status': 'error',
                'message': 'Email or Employee ID and password are required'
            }), 400

        teachers_ref = db.collection('teachers')
        # Try querying by email
        teacher_query = teachers_ref.where('email', '==', data['loginId']).limit(1).stream()
        teacher_doc = None
        for doc in teacher_query:
            teacher_doc = doc
            break
        
        # If not found by email, try by employeeId
        if not teacher_doc:
            teacher_query = teachers_ref.where('employeeId', '==', data['loginId']).limit(1).stream()
            for doc in teacher_query:
                teacher_doc = doc
                break

        if not teacher_doc:
            logger.error(f"Teacher not found for loginId: {data['loginId']}")
            return jsonify({
                'status': 'error',
                'message': 'Email or Employee ID not found'
            }), 404

        teacher_data = teacher_doc.to_dict()
        email = teacher_data['email']
        tid = teacher_doc.id
        logger.debug(f"Found teacher: {tid}, email: {email}")

        # Verify teacher role in Firebase Auth custom claims
        try:
            user = auth.get_user(tid)
            if not (hasattr(user, 'custom_claims') and user.custom_claims and user.custom_claims.get('role') == 'teacher'):
                logger.error(f"User {tid} does not have teacher role")
                return jsonify({
                    'status': 'error',
                    'message': 'Access denied: Not a teacher account'
                }), 403
        except auth.UserNotFoundError:
            logger.error(f"Firebase Auth user not found for teacher_id: {tid}")
            return jsonify({
                'status': 'error',
                'message': 'User account not found in authentication system'
            }), 404

        # Authenticate with Firebase
        url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={API_KEY}"
        payload = {
            "email": email,
            "password": data['password'],
            "returnSecureToken": True
        }
        response = requests.post(url, json=payload)
        
        if response.status_code == 200:
            session['teacher_id'] = tid
            logger.debug(f"Teacher login successful for teacher_id: {tid}")
            return jsonify({
                'status': 'success',
                'message': 'Login successful! Redirecting to teacher portal...'
            }), 200
        else:
            error_info = response.json().get('error', {})
            message = error_info.get('message', 'Incorrect password')
            if 'INVALID_PASSWORD' in message:
                message = 'Incorrect password'
            elif 'USER_NOT_FOUND' in message:
                message = 'User not found'
            elif 'INVALID_LOGIN_CREDENTIALS' in message:
                message = 'Invalid login credentials'
            logger.error(f"Teacher login failed: {message}")
            return jsonify({
                'status': 'error',
                'message': message
            }), 401

    except Exception as e:
        logger.error(f"Error during teacher login: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# API endpoint to send password reset email
@app.route('/api/send_password_reset', methods=['POST'])
@login_required
def send_password_reset():
    try:
        user_id = session.get('user_id')
        if not user_id:
            logger.warning("No user_id in session for send_password_reset")
            return jsonify({
                'status': 'error',
                'message': 'User not logged in'
            }), 401

        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            logger.error(f"User data not found for user_id: {user_id}")
            return jsonify({
                'status': 'error',
                'message': 'User data not found'
            }), 404

        user_data = user_doc.to_dict()
        email = user_data.get('email')
        if not email:
            logger.error(f"Email not found for user_id: {user_id}")
            return jsonify({
                'status': 'error',
                'message': 'Email not found for user'
            }), 404

        url = f"https://identitytoolkit.googleapis.com/v1/accounts:sendOobCode?key={API_KEY}"
        payload = {
            "requestType": "PASSWORD_RESET",
            "email": email
        }
        response = requests.post(url, json=payload)
        
        if response.status_code == 200:
            logger.debug(f"Password reset email sent to: {email}")
            return jsonify({
                'status': 'success',
                'message': 'Password reset email sent successfully'
            }), 200
        else:
            error_info = response.json().get('error', {})
            message = error_info.get('message', 'Failed to send password reset email')
            logger.error(f"Failed to send password reset email: {message}")
            return jsonify({
                'status': 'error',
                'message': message
            }), 400

    except Exception as e:
        logger.error(f"Error sending password reset email: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# API endpoint to sign out
@app.route('/api/sign_out', methods=['POST'])
def sign_out():
    try:
        session.pop('user_id', None)
        session.pop('teacher_id', None)
        logger.debug("User signed out successfully")
        return jsonify({
            'status': 'success',
            'message': 'Signed out successfully'
        }), 200
    except Exception as e:
        logger.error(f"Error signing out: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# API endpoint to delete account
@app.route('/api/delete_account', methods=['POST'])
@login_required
def delete_account():
    try:
        user_id = session.get('user_id')
        if not user_id:
            logger.warning("No user_id in session for delete_account")
            return jsonify({
                'status': 'error',
                'message': 'User not logged in'
            }), 401

        user_ref = db.collection('users').document(user_id)
        user_ref.delete()
        logger.debug(f"Deleted Firestore data for user_id: {user_id}")

        auth.delete_user(user_id)
        logger.debug(f"Deleted Firebase Auth user: {user_id}")

        session.pop('user_id', None)

        return jsonify({
            'status': 'success',
            'message': 'Account deleted successfully'
        }), 200

    except Exception as e:
        logger.error(f"Error deleting account: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }), 500

# API endpoint to get all students
@app.route('/api/get_all_students', methods=['GET'])
@login_required
def get_all_students():
    try:
        users = db.collection('users').stream()
        students = []
        for user in users:
            data = user.to_dict()
            students.append({
                'regNo': data.get('regNo', ''),
                'name': data.get('name', '')
            })
        students.sort(key=lambda x: x['regNo'])
        logger.debug(f"Fetched {len(students)} students")
        return jsonify({'status': 'success', 'students': students}), 200
    except Exception as e:
        logger.error(f"Error getting all students: {str(e)}")
        return jsonify({'status': 'error', 'message': 'An error occurred'}), 500

# API endpoint to get registered students (with face data)
@app.route('/api/get_registered_students', methods=['GET'])
@login_required
def get_registered_students():
    try:
        users = db.collection('users').where('face_encoding', '!=', None).stream()
        students = []
        for user in users:
            data = user.to_dict()
            face_image = data.get('face_image')
            if face_image:
                face_image = f"data:image/jpeg;base64,{face_image}"
            students.append({
                'regNo': data.get('regNo', ''),
                'name': data.get('name', ''),
                'face_image': face_image
            })
        students.sort(key=lambda x: x['regNo'])
        logger.debug(f"Fetched {len(students)} registered students")
        return jsonify({'status': 'success', 'students': students}), 200
    except Exception as e:
        logger.error(f"Error getting registered students: {str(e)}")
        return jsonify({'status': 'error', 'message': 'An error occurred'}), 500

# API endpoint to register face
@app.route('/api/register_face', methods=['POST'])
@login_required
def register_face():
    try:
        data = request.get_json()
        regNo = data.get('regNo')
        image = data.get('image')
        if not regNo or not image:
            logger.error("Missing regNo or image in register_face request")
            return jsonify({'status': 'error', 'message': 'Missing regNo or image'}), 400
        encoding = get_face_encoding(image)
        users_ref = db.collection('users')
        query = users_ref.where('regNo', '==', regNo).limit(1).stream()
        user_doc = next(query, None)
        if not user_doc:
            logger.error(f"Student not found: {regNo}")
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        image_base64 = image.split(',')[1] if ',' in image else image
        user_doc.reference.update({
            'face_encoding': encoding.tolist(),
            'face_image': image_base64,
            'updated_at': datetime.datetime.now()
        })
        logger.debug(f"Registered face for regNo: {regNo}")
        return jsonify({'status': 'success', 'message': 'Face registered successfully'}), 200
    except ValueError as ve:
        logger.error(f"ValueError in register_face: {str(ve)}")
        return jsonify({'status': 'error', 'message': str(ve)}), 400
    except Exception as e:
        logger.error(f"Error registering face: {str(e)}")
        return jsonify({'status': 'error', 'message': 'An error occurred'}), 500

# API endpoint to recognize face and mark attendance
@app.route('/api/recognize_face', methods=['POST'])
@login_required
def recognize_face():
    try:
        data = request.get_json()
        image = data.get('image')
        if not image:
            logger.error("Missing image in recognize_face request")
            return jsonify({'status': 'error', 'message': 'Missing image'}), 400
        captured_encoding = get_face_encoding(image)
        users = db.collection('users').where('face_encoding', '!=', None).stream()
        encodings = []
        students = []
        for user in users:
            user_data = user.to_dict()
            encodings.append(np.array(user_data['face_encoding']))
            students.append({
                'uid': user.id,
                'regNo': user_data.get('regNo', ''),
                'name': user_data.get('name', '')
            })
        if not encodings:
            logger.error("No registered faces found")
            raise ValueError("No registered faces found. Please register students first.")
        distances = face_recognition.face_distance(encodings, captured_encoding)
        min_distance_index = np.argmin(distances)
        min_distance = distances[min_distance_index]
        if min_distance > 0.5:
            logger.error("No matching face found")
            raise ValueError("No matching face found")
        matched_student = students[min_distance_index]
        today = datetime.date.today().isoformat()
        attendance_ref = db.collection('attendance').document(today)
        attendance_data = attendance_ref.get().to_dict() or {}
        if matched_student['regNo'] in attendance_data:
            logger.error(f"Attendance already marked for {matched_student['regNo']}")
            return jsonify({'status': 'error', 'message': 'Attendance already marked for today'}), 400
        attendance_data[matched_student['regNo']] = {
            'status': 'Present',
            'name': matched_student['name'],
            'timestamp': datetime.datetime.now()
        }
        attendance_ref.set(attendance_data)
        logger.debug(f"Marked attendance for {matched_student['regNo']}")
        return jsonify({
            'status': 'success',
            'studentId': matched_student['regNo'],
            'name': matched_student['name']
        }), 200
    except ValueError as ve:
        logger.error(f"ValueError in recognize_face: {str(ve)}")
        return jsonify({'status': 'error', 'message': str(ve)}), 400
    except Exception as e:
        logger.error(f"Error recognizing face: {str(e)}")
        return jsonify({'status': 'error', 'message': 'An error occurred'}), 500

# API endpoint to get today's attendance
@app.route('/api/get_today_attendance', methods=['GET'])
@login_required
def get_today_attendance():
    try:
        today = datetime.date.today().isoformat()
        attendance_ref = db.collection('attendance').document(today)
        attendance_data = attendance_ref.get().to_dict() or {}
        logger.debug(f"Fetched today's attendance: {attendance_data}")
        return jsonify({'status': 'success', 'attendance': attendance_data}), 200
    except Exception as e:
        logger.error(f"Error getting today's attendance: {str(e)}")
        return jsonify({'status': 'error', 'message': 'An error occurred'}), 500

# API endpoint to delete face data
@app.route('/api/delete_face', methods=['POST'])
@login_required
def delete_face():
    try:
        data = request.get_json()
        regNo = data.get('regNo')
        if not regNo:
            logger.error("Missing regNo in delete_face request")
            return jsonify({'status': 'error', 'message': 'Missing regNo'}), 400
        users_ref = db.collection('users')
        query = users_ref.where('regNo', '==', regNo).limit(1).stream()
        user_doc = next(query, None)
        if not user_doc:
            logger.error(f"Student not found: {regNo}")
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        user_doc.reference.update({
            'face_encoding': firestore.DELETE_FIELD,
            'face_image': firestore.DELETE_FIELD,
            'updated_at': datetime.datetime.now()
        })
        logger.debug(f"Deleted face data for regNo: {regNo}")
        return jsonify({'status': 'success', 'message': 'Face data deleted successfully'}), 200
    except Exception as e:
        logger.error(f"Error deleting face data: {str(e)}")
        return jsonify({'status': 'error', 'message': 'An error occurred'}), 500

# API endpoint to clear all face data
@app.route('/api/clear_all_faces', methods=['POST'])
@login_required
def clear_all_faces():
    try:
        users = db.collection('users').stream()
        batch = db.batch()
        for user in users:
            batch.update(user.reference, {
                'face_encoding': firestore.DELETE_FIELD,
                'face_image': firestore.DELETE_FIELD,
                'updated_at': datetime.datetime.now()
            })
        batch.commit()
        logger.debug("Cleared all face data")
        return jsonify({'status': 'success', 'message': 'All face data cleared'}), 200
    except Exception as e:
        logger.error(f"Error clearing all faces: {str(e)}")
        return jsonify({'status': 'error', 'message': 'An error occurred'}), 500

# API endpoint to reset today's attendance
@app.route('/api/reset_today_attendance', methods=['POST'])
@login_required
def reset_today_attendance():
    try:
        today = datetime.date.today().isoformat()
        db.collection('attendance').document(today).delete()
        logger.debug("Reset today's attendance")
        return jsonify({'status': 'success', 'message': 'Today\'s attendance reset'}), 200
    except Exception as e:
        logger.error(f"Error resetting attendance: {str(e)}")
        return jsonify({'status': 'error', 'message': 'An error occurred'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)